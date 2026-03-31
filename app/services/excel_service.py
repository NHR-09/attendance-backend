"""
Excel export service — Generates .xlsx attendance reports using openpyxl.
"""
import io
from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_attendance_excel(records: list[dict], report_date: date | None = None) -> io.BytesIO:
    """Generate a styled attendance Excel workbook.
    
    Each record dict should contain:
        employee_name, employee_id, confidence_score, date,
        check_in_time, check_out_time, method, status, notes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ── Styling ─────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    status_fills = {
        "present": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
        "absent": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
        "not_checked_out": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
        "exception": PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid"),
    }

    # ── Title row ───────────────────────────────────────────────
    title = f"Attendance Report — {report_date or date.today()}"
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # ── Header row ──────────────────────────────────────────────
    headers = [
        "Employee Name", "Employee ID", "Confidence %",
        "Date", "Check-In Time", "Check-Out Time",
        "Method", "Status", "Notes"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data rows ───────────────────────────────────────────────
    for row_idx, rec in enumerate(records, 4):
        values = [
            rec.get("employee_name", ""),
            rec.get("employee_id", ""),
            rec.get("confidence_score", ""),
            str(rec.get("date", "")),
            str(rec.get("check_in_time", "")),
            str(rec.get("check_out_time", "") or "—"),
            rec.get("method", ""),
            rec.get("status", ""),
            rec.get("notes", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # colour-code the status column
        status_val = rec.get("status", "").lower()
        status_cell = ws.cell(row=row_idx, column=8)
        if status_val in status_fills:
            status_cell.fill = status_fills[status_val]
            status_cell.font = Font(color="FFFFFF", bold=True)

    # ── Column widths ───────────────────────────────────────────
    widths = [22, 14, 14, 14, 18, 18, 12, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
